import openpyxl
import uuid
import datetime
from database import SessionLocal
from models import Branch, Class, FeeHead, FeeStructure, Student, Parent, Staff, User
from passlib.context import CryptContext

db = SessionLocal()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def run_import():
    from sqlalchemy import text
    print("Wiping existing test data...")
    db.execute(text("TRUNCATE TABLE branches CASCADE;"))
    db.execute(text("TRUNCATE TABLE classes CASCADE;"))
    db.execute(text("TRUNCATE TABLE staff CASCADE;"))
    db.execute(text("TRUNCATE TABLE users CASCADE;"))
    db.execute(text("TRUNCATE TABLE parents CASCADE;"))
    db.execute(text("TRUNCATE TABLE students CASCADE;"))
    db.commit()

    print("Inserting Branches...")
    b1_id = str(uuid.uuid4())
    b_akshaya_id = str(uuid.uuid4())
    db.add(Branch(id=b1_id, name='SPKOVIL BRANCH1', code='SPK1', address='SPKOVIL'))
    db.add(Branch(id=b_akshaya_id, name='AKSHAYA', code='AKS', address='AKSHAYA'))
    db.commit()

    print("Recreating Super Admin...")
    admin_id = str(uuid.uuid4())
    admin_user = User(
        id=admin_id,
        username='principal_admin',
        password_hash=pwd_context.hash('password123'),
        role='ADMIN'
    )
    db.add(admin_user)
    db.commit()

    branches = {'SPKOVIL BRANCH1': b1_id, 'AKSHAYA': b_akshaya_id}

    # STAFF
    print("Importing Staff...")
    wb_staff = openpyxl.load_workbook('/app/STAFF DATA.xlsx', read_only=True, data_only=True)
    sheet = wb_staff['Sheet1']
    staff_count = 0
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row[1]: continue
        name = str(row[1]).strip()
        b_name = str(row[2]).strip().upper() if len(row) > 2 and row[2] else 'AKSHAYA'
        qual = str(row[3]) if len(row) > 3 and row[3] else None
        role = str(row[4]) if len(row) > 4 and row[4] else 'Staff'
        doj_val = row[5] if len(row) > 5 else None
        doj = doj_val.date() if isinstance(doj_val, datetime.datetime) else datetime.date.today()
        email = str(row[6]) if len(row) > 6 and row[6] else None
        phone = str(row[7]) if len(row) > 7 and row[7] else None
        addr = str(row[8]) if len(row) > 8 and row[8] else None
        
        branch_id = branches.get(b_name, b_akshaya_id)
        
        staff = Staff(
            id=str(uuid.uuid4()),
            name=name,
            role=role,
            phone=phone,
            email=email,
            qualification=qual,
            address=addr,
            joining_date=doj,
            monthly_salary=10000.0,
            branch_id=branch_id
        )
        db.add(staff)
        staff_count += 1
    db.commit()

    # CLASSES AND STUDENTS
    print("Importing Students & Parents...")
    wb_students = openpyxl.load_workbook('/app/SPKOVIL BRANCH1 2026-27.xlsx', read_only=True, data_only=True)
    student_count = 0
    parent_count = 0
    class_count = 0
    parent_cache = {}

    student_cache = set()

    for sheet_name in wb_students.sheetnames:
        class_name = sheet_name.strip()
        cls_id = str(uuid.uuid4())
        db.add(Class(id=cls_id, name=class_name))
        class_count += 1
        db.commit()

        sheet = wb_students[sheet_name]
        
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if len(row) < 3 or not row[0] or not row[1] or not row[2]:
                continue
                
            adm_no = str(row[1]).strip()
            if 'adm' in adm_no.lower():
                continue
                
            if adm_no in student_cache:
                continue
            student_cache.add(adm_no)
            name = str(row[2]).strip()
            father = str(row[3]).strip() if len(row) > 3 and row[3] else "Father"
            mother = str(row[4]).strip() if len(row) > 4 and row[4] else "Mother"
            phone = str(row[5]).strip() if len(row) > 5 and row[5] else "0000000000"
            dob_str = str(row[6]).strip() if len(row) > 6 and row[6] else None
            bg = str(row[7]).strip() if len(row) > 7 and row[7] else None
            
            if phone not in parent_cache:
                p_id = str(uuid.uuid4())
                parent = Parent(id=p_id, father_name=father, mother_name=mother, primary_phone=phone)
                db.add(parent)
                parent_cache[phone] = p_id
                parent_count += 1
            else:
                p_id = parent_cache[phone]

            stu = Student(
                id=str(uuid.uuid4()),
                roll_no=adm_no,
                name=name,
                class_id=cls_id,
                parent_id=p_id,
                blood_group=bg,
                status="ACTIVE",
                branch_id=b1_id
            )
            db.add(stu)
            student_count += 1
    db.commit()

    print(f"Import Complete!")
    print(f"Branches: 2")
    print(f"Classes: {class_count}")
    print(f"Staff: {staff_count}")
    print(f"Parents: {parent_count}")
    print(f"Students: {student_count}")

if __name__ == '__main__':
    run_import()
