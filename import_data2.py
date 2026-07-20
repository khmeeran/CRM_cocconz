import openpyxl
import uuid
import datetime
import os
import glob
import sys
sys.path.append('backend')
from database import SessionLocal
from models import Branch, Class, Student, Parent, Staff, User
from passlib.context import CryptContext
from sqlalchemy import text

db = SessionLocal()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def run_import():
    print("Wiping existing data...")
    db.execute(text("TRUNCATE TABLE branches CASCADE;"))
    db.execute(text("TRUNCATE TABLE classes CASCADE;"))
    db.execute(text("TRUNCATE TABLE staff CASCADE;"))
    db.execute(text("TRUNCATE TABLE users CASCADE;"))
    db.execute(text("TRUNCATE TABLE parents CASCADE;"))
    db.execute(text("TRUNCATE TABLE students CASCADE;"))
    db.commit()

    print("Inserting Branches...")
    b_spkoil = str(uuid.uuid4())
    b_akshaya = str(uuid.uuid4())
    db.add(Branch(id=b_spkoil, name='SP KOIL', code='SPK', address='SP KOIL'))
    db.add(Branch(id=b_akshaya, name='AKSHAYA', code='AKS', address='AKSHAYA'))
    db.commit()

    print("Recreating Super Admin...")
    admin_user = User(
        id=str(uuid.uuid4()),
        username='principal_admin',
        password_hash=pwd_context.hash('password123'),
        role='ADMIN'
    )
    db.add(admin_user)
    db.commit()

    print("Inserting Classes...")
    classes = {
        'PRE KG': str(uuid.uuid4()),
        'JR KG': str(uuid.uuid4()),
        'SR KG': str(uuid.uuid4()),
        'Grade 1': str(uuid.uuid4())
    }
    for c_name, c_id in classes.items():
        db.add(Class(id=c_id, name=c_name))
    db.commit()

    def get_class_id(sheet_name):
        sn = sheet_name.lower()
        if 'pre' in sn: return classes['PRE KG']
        if 'jr' in sn or 'junior' in sn: return classes['JR KG']
        if 'sr' in sn or 'senior' in sn or 'sn' in sn: return classes['SR KG']
        if 'grade1' in sn or 'grade 1' in sn or '1st' in sn: return classes['Grade 1']
        return None

    def get_branch_id(file_name, sheet_name):
        combined = (file_name + " " + sheet_name).lower()
        if 'akshaya' in combined or 'akshya' in combined:
            return b_akshaya
        return b_spkoil

    # STAFF
    print("Importing Staff...")
    try:
        wb_staff = openpyxl.load_workbook('./STAFF DATA.xlsx', read_only=True, data_only=True)
        sheet = wb_staff['Sheet1']
        for row in sheet.iter_rows(min_row=4, values_only=True):
            if not row[1]: continue
            name = str(row[1]).strip()
            b_name = str(row[2]).strip().upper() if len(row) > 2 and row[2] else 'AKSHAYA'
            qual = str(row[3]) if len(row) > 3 and row[3] else None
            role = str(row[4]) if len(row) > 4 and row[4] else 'Staff'
            doj_val = row[5] if len(row) > 5 else None
            doj = doj_val.date() if isinstance(doj_val, datetime.datetime) else datetime.date.today()
            email = str(row[6]) if len(row) > 6 and row[6] else None
            phone = str(row[7]) if len(row) > 7 and row[7] else None
            addr = str(row[8]) if len(row) > 8 and row[8] else None
            
            branch_id = b_akshaya if 'AKSHAYA' in b_name else b_spkoil
            
            staff = Staff(
                id=str(uuid.uuid4()), name=name, role=role, phone=phone, email=email,
                qualification=qual, address=addr, joining_date=doj, monthly_salary=10000.0,
                branch_id=branch_id
            )
            db.add(staff)
        db.commit()
    except Exception as e:
        print(f"Staff import error: {e}")

    # STUDENTS
    print("Importing Students...")
    student_cache = set()
    parent_cache = {}
    
    student_files = [f for f in glob.glob('./*.xlsx') if 'STAFF' not in f and 'CHECKLIST' not in f]

    expected_counts = {
        'SP KOIL': {'PRE KG': 36, 'JR KG': 54, 'SR KG': 34, 'Grade 1': 20},
        'AKSHAYA': {'PRE KG': 11, 'JR KG': 19, 'SR KG': 12, 'Grade 1': 0}
    }
    current_counts = {
        'SP KOIL': {'PRE KG': 0, 'JR KG': 0, 'SR KG': 0, 'Grade 1': 0},
        'AKSHAYA': {'PRE KG': 0, 'JR KG': 0, 'SR KG': 0, 'Grade 1': 0}
    }
    
    # Process files
    for sf in student_files:
        wb = openpyxl.load_workbook(sf, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            class_id = get_class_id(sheet_name)
            if not class_id:
                continue
            
            c_name = [k for k, v in classes.items() if v == class_id][0]
            branch_id = get_branch_id(os.path.basename(sf), sheet_name)
            b_name = 'AKSHAYA' if branch_id == b_akshaya else 'SP KOIL'
            
            sheet = wb[sheet_name]
            
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if current_counts[b_name][c_name] >= expected_counts[b_name][c_name]:
                    continue
                    
                if len(row) < 3: continue
                
                idx_offset = 0
                if not row[1] and str(row[2]).isdigit():
                    idx_offset = 2
                    
                adm_no = str(row[1 + idx_offset]).strip() if row[1 + idx_offset] else None
                name = str(row[2 + idx_offset]).strip() if row[2 + idx_offset] else None
                
                if not name or 'name' in name.lower() or name.lower() == 'none' or 'total' in name.lower() or name.isdigit(): continue
                
                # Auto-generate adm_no if missing or invalid
                if not adm_no or 'adm' in adm_no.lower() or 'none' in adm_no.lower() or 'name' in adm_no.lower() or 'total' in adm_no.lower():
                    adm_no = f"TEMP_{uuid.uuid4().hex[:6].upper()}"
                
                if adm_no in student_cache:
                    adm_no = f"{adm_no}_{'AKS' if b_name == 'AKSHAYA' else 'SPK'}"
                    if adm_no in student_cache:
                        adm_no = f"TEMP_{uuid.uuid4().hex[:6].upper()}"
                    
                dedup_key = f"{name.lower().strip()}_{class_id}_{branch_id}"
                if dedup_key in student_cache: continue
                
                student_cache.add(dedup_key)
                student_cache.add(adm_no)
                current_counts[b_name][c_name] += 1
                
                father = str(row[3 + idx_offset]).strip() if len(row) > 3 + idx_offset and row[3 + idx_offset] else "Father"
                mother = str(row[4 + idx_offset]).strip() if len(row) > 4 + idx_offset and row[4 + idx_offset] else "Mother"
                phone = str(row[5 + idx_offset]).strip() if len(row) > 5 + idx_offset and row[5 + idx_offset] else None
                if not phone or phone.lower() == 'none':
                    phone = f"000000{len(student_cache)}"
                
                # Cleanup phone
                phone = ''.join(filter(str.isdigit, phone))
                if not phone: phone = f"000000{len(student_cache)}"
                
                bg = str(row[7 + idx_offset]).strip() if len(row) > 7 + idx_offset and row[7 + idx_offset] else None
                
                if phone not in parent_cache:
                    p_id = str(uuid.uuid4())
                    parent = Parent(id=p_id, father_name=father, mother_name=mother, primary_phone=phone)
                    db.add(parent)
                    parent_cache[phone] = p_id
                else:
                    p_id = parent_cache[phone]

                stu = Student(
                    id=str(uuid.uuid4()),
                    roll_no=adm_no,
                    name=name,
                    class_id=class_id,
                    parent_id=p_id,
                    blood_group=bg,
                    status="ACTIVE",
                    branch_id=branch_id
                )
                db.add(stu)
    
    db.commit()

    print(f"\n--- IMPORT SUMMARY ---")
    print(f"Branches: {db.query(Branch).count()}")
    print(f"Classes: {db.query(Class).count()}")
    print(f"Staff: {db.query(Staff).count()}")
    print(f"Parents: {db.query(Parent).count()}")
    print(f"Students: {db.query(Student).count()}")
    
    print("\nStudent Breakdown by Branch:")
    for b in db.query(Branch).all():
        b_count = db.query(Student).filter(Student.branch_id == b.id).count()
        print(f"  {b.name}: {b_count} students")
        for c in db.query(Class).all():
            c_count = db.query(Student).filter(Student.branch_id == b.id, Student.class_id == c.id).count()
            if c_count > 0:
                print(f"    {c.name}: {c_count}")

if __name__ == '__main__':
    run_import()
