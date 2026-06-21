import logging
from typing import List, Dict, Any
import os
from fpdf import FPDF
import openpyxl
from datetime import datetime

logger = logging.getLogger(__name__)

class NotificationService:
    @staticmethod
    def send_sms(phone_numbers: List[str], message: str):
        # In a real production system, integrate with Twilio or AWS SNS
        for phone in phone_numbers:
            logger.info(f"Mock SMS sent to {phone}: {message}")

    @staticmethod
    def send_whatsapp(phone_numbers: List[str], message: str):
        # In a real production system, integrate with WhatsApp Business API
        for phone in phone_numbers:
            logger.info(f"Mock WhatsApp sent to {phone}: {message}")

# Background worker wrapper
def broadcast_worker(target_phones: List[str], message: str, channels: List[str]):
    try:
        if "SMS" in channels:
            NotificationService.send_sms(target_phones, message)
        if "WHATSAPP" in channels:
            NotificationService.send_whatsapp(target_phones, message)
    except Exception as e:
        logger.error(f"Failed to process broadcast: {e}")

class ExportService:
    @staticmethod
    def generate_excel(data: List[Dict[str, Any]], filename: str) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export Data"

        if len(data) > 0:
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([str(row.get(h, "")) for h in headers])

        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, f"{filename}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
        wb.save(filepath)
        return filepath

    @staticmethod
    def generate_pdf(data: List[Dict[str, Any]], title: str, filename: str) -> str:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(200, 10, txt=title, ln=True, align='C')
        pdf.ln(10)
        
        pdf.set_font("Arial", size=10)
        
        if len(data) > 0:
            headers = list(data[0].keys())
            col_width = 190 / max(len(headers), 1)
            
            # Header
            pdf.set_font("Arial", 'B', 10)
            for h in headers:
                pdf.cell(col_width, 10, str(h), border=1)
            pdf.ln()
            
            # Rows
            pdf.set_font("Arial", size=10)
            for row in data:
                for h in headers:
                    pdf.cell(col_width, 10, str(row.get(h, ""))[:20], border=1)
                pdf.ln()

        export_dir = os.path.join(os.path.dirname(__file__), "exports")
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, f"{filename}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")
        pdf.output(filepath)
        return filepath

    @staticmethod
    def generate_receipt_pdf(receipt_no: str, student_name: str, roll_no: str, branch: str, class_name: str, fee_head: str, amount: float, balance: float, date: str, payment_mode: str, collected_by: str) -> bytearray:
        pdf = FPDF()
        pdf.add_page()
        
        # Header
        pdf.set_font("Helvetica", 'B', 20)
        pdf.cell(190, 10, text="COCOONZ PRESCHOOL", ln=True, align='C')
        pdf.set_font("Helvetica", '', 10)
        pdf.cell(190, 6, text="Official Fee Receipt", ln=True, align='C')
        pdf.ln(10)
        
        def add_row(lbl, val):
            pdf.set_font("Helvetica", 'B', 12)
            pdf.cell(50, 10, text=str(lbl), border=0)
            pdf.set_font("Helvetica", '', 12)
            pdf.cell(140, 10, text=str(val), border=0, ln=True)

        add_row("Receipt No:", receipt_no)
        add_row("Date:", date)
        add_row("Student Name:", student_name)
        add_row("Admission No:", roll_no)
        add_row("Branch:", branch)
        add_row("Class:", class_name)
        add_row("Fee Head:", fee_head)
        add_row("Payment Mode:", payment_mode)
        add_row("Collected By:", collected_by)
        
        pdf.ln(5)
        
        # Amount Box
        pdf.set_font("Helvetica", 'B', 14)
        pdf.cell(50, 15, text="Amount Paid:", border=1)
        pdf.set_font("Helvetica", 'B', 16)
        pdf.cell(140, 15, text=f"INR {amount}", border=1, ln=True)
        
        pdf.set_font("Helvetica", 'B', 12)
        pdf.cell(50, 15, text="Balance Remaining:", border=1)
        pdf.set_font("Helvetica", '', 14)
        pdf.cell(140, 15, text=f"INR {balance}", border=1, ln=True)
        
        pdf.ln(10)
        pdf.set_font("Helvetica", 'I', 10)
        pdf.cell(190, 10, text="This is a computer generated receipt.", ln=True, align='C')
        
        return pdf.output()

